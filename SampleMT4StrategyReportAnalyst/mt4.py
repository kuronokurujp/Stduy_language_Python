from pathlib import Path
import statistics
import bs4  # ライブラリbs4をインポートする
import openpyxl


def __change_name(src_name: str, change_name_dict: dict[str, str]):
    if change_name_dict == None:
        return src_name

    if src_name in change_name_dict:
        return change_name_dict[src_name]

    return src_name


def __change_param_name(src: str, param_name_dict: dict[str, str]):
    if param_name_dict == None:
        return src

    conv_str: str = src

    for k, v in param_name_dict.items():
        conv_str = conv_str.replace(k, v)

    return conv_str


class BaseDataModel(object):
    def __init__(self) -> None:
        # パス
        self.no: float = 0.0
        # 損益
        self.prfoit_lost: float = 0.0
        # 総取引数
        self.total_number_transactions: float = 0.0
        # プロフィットファクタ
        self.profit_factor: float = 0.0
        # 期待利得
        self.expected_gain: float = 0.0
        # ドローダウン $
        self.drawdown_d: float = 0.0
        # ドローダウン %
        self.drawdown_p: float = 0.0


class DataModel(BaseDataModel):
    def __init__(self, curreny_name: str, html_list: list = None) -> None:
        super().__init__()

        if not html_list is None:
            count: int = 0
            for v in vars(self).items():
                e_soup: bs4.BeautifulStoneSoup = bs4.BeautifulSoup(
                    str(html_list[count]), "html.parser"
                )
                val_list = list(e_soup.find("td"))

                key: str = str(v[0])
                value: float = float(val_list[0])
                self.__dict__[key] = value

                count = count + 1

            # パラメータ取得
            e_soup: bs4.BeautifulStoneSoup = bs4.BeautifulSoup(
                str(html_list[0]), "html.parser"
            )
            val_list = list(e_soup.find_all("td"))
            self.params = val_list[0].attrs["title"]
        else:
            self.params = ""

        self.curreny_name = curreny_name


class ParamPairModel(object):
    def __init__(self, model_list: list) -> None:
        # パス
        self.no: float = model_list[0].no
        self.params: str = model_list[0].params
        self.model_list: list[DataModel] = model_list

        self.params_dict: dict[str, list] = dict()
        params_list: list = self.params.split(";")
        for p in params_list:
            items = p.split("=")
            if len(items) == 2:
                try:
                    name: str = items[0]
                    name = name.lstrip()
                    name = name.rstrip()

                    f: float = float(items[1])
                    self.params_dict[name] = f
                except Exception as ex:
                    print(
                        "error change float or int {} = {} type={}",
                        items[0],
                        items[1],
                        type(items[1]),
                    )
                    raise ex

    def get_currency_count(self) -> int:
        return len(self.model_list)

    def get_pair_curreny_name(self):
        comment: str = ""
        for model in self.model_list:
            comment = comment + "{},".format(model.curreny_name)
        return comment

    def get_comment(self, type_name: str) -> str:
        comment: str = ""
        for model in self.model_list:
            comment = comment + "{}: {}".format(
                model.curreny_name, getattr(model, type_name)
            )
            comment = comment + "\n"
        return comment

    def get_comment_by_param(self) -> str:
        return ""
        comment: str = ""
        for k, v in self.params_dict.items():
            add_comment: str = "{}={}".format(k, v)
            if comment == "":
                comment = add_comment
            else:
                comment = "{}\n{}".format(comment, add_comment)

        return comment

    def get_avg_val(self, type_name: str) -> float:
        match type_name:
            case "prfoit_lost":
                # 損益
                return statistics.median([val.prfoit_lost for val in self.model_list])
            case "total_number_transactions":
                # 総取引数
                return statistics.median(
                    [val.total_number_transactions for val in self.model_list]
                )
            case "profit_factor":
                # プロフィットファクタ
                return statistics.median([val.profit_factor for val in self.model_list])
            case "expected_gain":
                # 期待利得
                return statistics.median([val.expected_gain for val in self.model_list])
            case "drawdown_d":
                # ドローダウン $
                return statistics.median([val.drawdown_d for val in self.model_list])
            case "drawdown_p":
                # ドローダウン %
                return statistics.median([val.drawdown_p for val in self.model_list])

        return 0.0


def load_mt4_html(filename_fullpath: str) -> bs4.BeautifulStoneSoup:
    file_path = Path(filename_fullpath)
    # win版のMT4で作成したhtmlファイルの文字コードはshift-jisになる
    with open(str(file_path), encoding="shift-jis") as f:
        soup: bs4.BeautifulStoneSoup = bs4.BeautifulSoup(f, "html.parser")

        return soup


def create_currency_name(filename_fullpath: str) -> str:
    file_path = Path(filename_fullpath)
    return file_path.stem.split("_")[0]


def create_model_list(
    currenty_name: str, soup: bs4.BeautifulStoneSoup
) -> list[DataModel]:
    element_table = soup.find_all("table")

    # print(element_table)
    element_target_table = element_table[1]
    # print(element_target_table)

    element_target_table_soup = bs4.BeautifulStoneSoup = bs4.BeautifulSoup(
        str(element_target_table), "html.parser"
    )
    element_tr_list: list = element_target_table_soup.find_all("tr")
    del element_tr_list[0]

    data_model_list: list[DataModel] = list()
    for e in element_tr_list:
        e_soup: bs4.BeautifulStoneSoup = bs4.BeautifulSoup(str(e), "html.parser")
        td_list: list = e_soup.find_all("td")

        model = DataModel(curreny_name=currenty_name, html_list=td_list)
        # 取引回数がない場合は検証失敗しているデータなので無視
        if 0 < model.total_number_transactions:
            data_model_list.append(model)

    return data_model_list


def create_param_group_model_dict(
    model_dict: dict[str, list[DataModel]]
) -> dict[str, list[DataModel]]:
    group_model_dict: dict[str, list[DataModel]] = dict()

    count: int = 0
    for key, model_list in model_dict.items():
        # パラメータをペアとしたモデルを作成
        for model in model_list:
            if not model.params in group_model_dict:
                group_model_dict[model.params] = list()

            group_model_dict[model.params].append(model)
            if count < len(group_model_dict[model.params]):
                count = len(group_model_dict[model.params])

    return group_model_dict


# floatをセル書き込み
def write_cell_float(sheet, cell_name: str, val: float):
    sheet[cell_name] = val
    # 小数点以下を2桁まで表示
    sheet[cell_name].number_format = "0.00"


def write_cell_ea_param(
    sheet,
    cell_name: str,
    model: ParamPairModel,
    param_name: str,
    jp_name_dict: dict[str, str] = None,
):
    from openpyxl.comments import Comment

    comment: Comment = None
    if param_name == "params":
        sheet[cell_name] = __change_param_name(
            model.params, param_name_dict=jp_name_dict
        )
        # コメントがあれば追加する
        if model.get_comment_by_param() != "":
            comment = Comment(model.get_comment_by_param(), "Comment Author")
            comment.height = 400
    else:
        write_cell_float(sheet, cell_name, model.get_avg_val(type_name=param_name))
        # sheet[cell_name] = model.get_avg_val(type_name=param_name)
        # # 小数点以下を2桁まで表示
        # sheet[cell_name].number_format = "0.00"
        # 各通貨の情報をコメントにする
        comment = Comment(model.get_comment(type_name=param_name), "Comment Author")
        comment.height = 200

    if comment is not None:
        sheet[cell_name].comment = comment


def output_xlsx(
    dir_fullpath: str,
    filename: str,
    pair_model_list: list[ParamPairModel],
    curreny_model_dict: dict[str, list[DataModel]],
    param_name_dict: dict[str, str] = None,
):
    file_path = Path(dir_fullpath) / "{}.xlsx".format(filename)
    file_path.unlink(True)

    # ブックを作成
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    # シート名は通貨のペア数を元に作る
    sheet = wb.create_sheet("概要", 0)

    # 成績表を作る
    curreny_result_header_dict: dict[str, str] = {
        "G": "通貨",
        "H": "利益を出したパターン数",
        "I": "損失を出したパターン数",
        "J": "利益も損失も出していないパターン数",
        "K": "損益の中央値",
        "L": "総取引数の中央値",
        "M": "プロフィットファクターの中央値",
        "N": "期待利得の中央値",
        "O": "ドローダウン$の中央値",
        "P": "ドローダウン%の中央値",
    }
    for key, value in curreny_result_header_dict.items():
        sheet["{}{}".format(key, 1)] = value

    from statistics import mean, median

    curreny_result_row: int = 2
    for key, model_list in curreny_model_dict.items():
        sheet["G" + str(curreny_result_row)] = key
        sheet["H" + str(curreny_result_row)] = len(
            list(filter(lambda x: 1 < x.profit_factor, model_list))
        )
        sheet["I" + str(curreny_result_row)] = len(
            list(filter(lambda x: x.profit_factor < 1, model_list))
        )
        sheet["J" + str(curreny_result_row)] = len(
            list(filter(lambda x: x.profit_factor == 1.0, model_list))
        )

        if 0 < len(model_list):
            write_cell_float(
                sheet,
                "K" + str(curreny_result_row),
                median([x.prfoit_lost for x in model_list]),
            )
            write_cell_float(
                sheet,
                "L" + str(curreny_result_row),
                median([x.total_number_transactions for x in model_list]),
            )
            write_cell_float(
                sheet,
                "M" + str(curreny_result_row),
                median([x.profit_factor for x in model_list]),
            )
            write_cell_float(
                sheet,
                "N" + str(curreny_result_row),
                median([x.expected_gain for x in model_list]),
            )
            write_cell_float(
                sheet,
                "O" + str(curreny_result_row),
                median([x.drawdown_d for x in model_list]),
            )
            write_cell_float(
                sheet,
                "P" + str(curreny_result_row),
                median([x.drawdown_p for x in model_list]),
            )
        else:
            sheet["F" + str(curreny_result_row)] = "データがない"

        curreny_result_row = curreny_result_row + 1

    # 各パラメータのリストを作る
    param_dict: dict[str, list[int | float]] = dict()
    for model_list in pair_model_list:
        for k, v in model_list.params_dict.items():
            if not k in param_dict:
                param_dict[k] = list()

            param_dict[k].append(v)

    row: int = 1
    # テストパラメータのパターンを記入
    for k, v in param_dict.items():
        row_str: str = str(row)
        # パラメータ名
        sheet["A" + row_str] = __change_name(k, change_name_dict=param_name_dict)
        # パラメータの最小
        sheet["B" + row_str] = min(v)

        # パラメータのステップ
        # 重複なしにする
        l_reverse_sorted = list(set(v))
        # 降順にソート
        l_reverse_sorted = sorted(l_reverse_sorted, reverse=True)
        if len(l_reverse_sorted) == 0:
            sheet["C" + row_str] = 0
        elif len(l_reverse_sorted) == 1:
            sheet["C" + row_str] = l_reverse_sorted[0]
        else:
            sheet["C" + row_str] = l_reverse_sorted[0] - l_reverse_sorted[1]

        # パラメータの最大
        sheet["D" + row_str] = max(v)

        row = row + 1

    # 全リストデータのシートを作成
    sheet = wb.create_sheet("全リスト", 1)
    write_sheet_test_data(
        row=row,
        sheet=sheet,
        pair_model_list=pair_model_list,
        param_name_dict=param_name_dict,
    )

    # 各通貨で共通しているパラメータの成績を別シートに記載
    max_pair_count: int = len(curreny_model_dict)
    for i in range(1, max_pair_count + 1):
        sheet = wb.create_sheet("{}通貨".format(i), 1 + i)
        # 共通しているパラメータの成績を抜き出してシートに書き込む
        now_pair_molde_list = filter(
            lambda x: x.get_currency_count() == i, pair_model_list
        )
        write_sheet_test_data(
            row=row,
            sheet=sheet,
            pair_model_list=now_pair_molde_list,
            param_name_dict=param_name_dict,
        )

    # 保存
    wb.save(str(file_path))


def write_sheet_test_data(
    row: int,
    sheet,
    pair_model_list: list[ParamPairModel],
    param_name_dict: dict[str, str],
):
    param_result_header_list: list = [
        "NO",
        "通貨数",
        "損益",
        "総取引数",
        "プロフィットファクタ",
        "期待利得",
        "ドローダウン $",
        "ドローダウン %",
        "パラメータ",
    ]

    name_col: int = 1
    h_row = row
    for name in param_result_header_list:
        sheet.cell(row=h_row, column=name_col, value=name)
        name_col = name_col + 1
    row = row + 1

    # 損益を降順で並べ替え
    new_pair_model_list: list[ParamPairModel] = sorted(
        pair_model_list,
        key=lambda model: model.get_avg_val("prfoit_lost"),
        reverse=True,
    )
    # データ数が１万を超えた場合はデータ整理する
    # データ数を1万までに抑える
    if 10000 <= len(new_pair_model_list):
        chk_pair_model_list: list[ParamPairModel] = list[ParamPairModel]()
        chk_pair_model_list.append(new_pair_model_list[0])

        duplicate_count: int = 0
        for i in range(1, 10000 - 1):
            prev_currency_type_name: str = new_pair_model_list[i - 1].get_pair_curreny_name()
            now_currency_type_name: str = new_pair_model_list[i].get_pair_curreny_name()

            # 同じ通貨が連続している場合は10個までは出力する
            # 同じ通貨が100とか1000とか連続しているデータは意味がないから
            if prev_currency_type_name == now_currency_type_name:
                if duplicate_count <= 10:
                    duplicate_count = duplicate_count + 1
                    chk_pair_model_list.append(new_pair_model_list[i])
            else:
                chk_pair_model_list.append(new_pair_model_list[i])
                duplicate_count = 0

        new_pair_model_list = chk_pair_model_list

    count: int = row
    for model_list in new_pair_model_list:
        count_str = str(count)
        sheet["A" + count_str] = str(count - 1)
        sheet["B" + count_str] = model_list.get_currency_count()
        write_cell_ea_param(
            sheet=sheet,
            cell_name="C" + count_str,
            model=model_list,
            param_name="prfoit_lost",
        )
        write_cell_ea_param(
            sheet=sheet,
            cell_name="D" + count_str,
            model=model_list,
            param_name="total_number_transactions",
        )
        write_cell_ea_param(
            sheet=sheet,
            cell_name="E" + count_str,
            model=model_list,
            param_name="profit_factor",
        )
        write_cell_ea_param(
            sheet=sheet,
            cell_name="F" + count_str,
            model=model_list,
            param_name="expected_gain",
        )
        write_cell_ea_param(
            sheet=sheet,
            cell_name="G" + count_str,
            model=model_list,
            param_name="drawdown_d",
        )
        write_cell_ea_param(
            sheet=sheet,
            cell_name="H" + count_str,
            model=model_list,
            param_name="drawdown_p",
        )
        write_cell_ea_param(
            sheet=sheet,
            cell_name="I" + count_str,
            model=model_list,
            param_name="params",
            jp_name_dict=param_name_dict,
        )

        count = count + 1
    # オートフィルタ範囲の設定
    sheet.auto_filter.ref = "{}:{}".format("A{}".format(h_row), "I{}".format(count))


def get_target_file_list(dir_filenamt: str, ignore_currenty_names: list) -> list:
    import glob

    file_fullpath: str = str(Path(dir_filenamt) / "*htm")

    if ignore_currenty_names is None:
        return [f for f in glob.glob(file_fullpath)]
    return [
        f for f in glob.glob(file_fullpath) if not Path(f).stem in ignore_currenty_names
    ]


# 通貨データファイルから各通貨のデータモデルリストを通貨名をキーとした辞書として作成
def create_currency_model_dict(path_filename_list: list) -> dict[str, list[DataModel]]:
    currenty_model_dict: dict[str, list[DataModel]] = dict()
    # 各通貨データファイルを解析して各通貨のモデルリストを作成
    for path_filename in path_filename_list:
        soup = load_mt4_html(path_filename)
        assert not soup is None

        # 辞書のキーを通貨名にして通貨毎のデータモデルリストを作成
        currenty_name: str = create_currency_name(path_filename)
        if not currenty_model_dict is currenty_name:
            currenty_model_dict[currenty_name] = list()

        # currenty_model_dict.extend(create_model_list(currenty_name=currenty_name, soup=soup))
        currenty_model_dict[currenty_name] = create_model_list(
            currenty_name=currenty_name, soup=soup
        )
        assert not currenty_model_dict[currenty_name] is None
        assert 0 < len(currenty_model_dict)

    return currenty_model_dict


# 通貨毎のモデルデータリストからパラメータをペアとしたモデルリストをコンバート
def conv_currenty_model_dict_to_param_pair_model_list(
    currenty_model_dict: dict[str, list[DataModel]]
) -> list[ParamPairModel]:
    # パラメータでまとめたモデル辞書作成
    param_group_model_dict: dict[list[DataModel]] = create_param_group_model_dict(
        model_dict=currenty_model_dict
    )
    assert not param_group_model_dict is None
    assert 0 < len(param_group_model_dict)

    # パラメータをペアとしたモデルリストを作成
    param_pair_model_list: list[ParamPairModel] = list()
    for key, value in param_group_model_dict.items():
        model: ParamPairModel = ParamPairModel(model_list=value)
        param_pair_model_list.append(model)

    return param_pair_model_list


def main():
    pass


if __name__ == "__main__":
    main()
