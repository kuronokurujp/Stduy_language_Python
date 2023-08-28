from fire import Fire

import datetime
import math
import openpyxl
import os
import pandas as pd
import time


# コマンド実行クラス
# csvファイル作成は標準関数を利用
# csv / pandasのモジュールだと数百万データを取り扱う場合は時間がかかるから
# https://zenn.dev/yagiyuki/articles/csv_performance_py
class Command(object):
    csv_write = None
    csv_file = None

    # セルの00:00:00のデータ
    excel_zero_date = datetime.datetime(1899, 12, 30, 0, 0, 0)

    # クラスのデストラクター
    # https://it-engineer-info.com/language/python/5689/
    def __del__(self):
        if self.csv_write:
            self.csv_write = None

        if self.csv_file:
            self.csv_file.close()
            self.csv_file = None

    # 作成した1分足のcsvファイルデータを分足などにリダクションしたcsvファイル作成
    def make_period_history(
        self,
        i_csv_fname_1min: str,
        o_dir: str,
        o_fname: str,
        minutes: int,
    ):
        # datetimeの列は型をdatetimeに変える
        # https://www.self-study-blog.com/dokugaku/python-pandas-csv-datetime-parse/
        origin_df = pd.read_csv(
            i_csv_fname_1min, encoding="utf-8", parse_dates=["day", "datetime"]
        )

        # 日付のリスト
        origin_day = origin_df.groupby("day")
        origin_day = origin_day.mean().index

        # 結果リスト
        result_list = []

        # 日付の繰り返し
        for day in origin_day:
            day_df = origin_df[origin_df["day"] == day]
            # 先頭の日付を開始時間にしている
            start_time = day_df.at[day_df.index[0], "datetime"]

            # r = range(math.ceil(len(day_df) / chart_term))
            b_loop: bool = True

            print("start parsing day({})".format(day))
            while b_loop:
                # for i in r:
                # print(f"\r\033[K({i}) / ({len(r)})", end="")

                end_time = start_time + datetime.timedelta(minutes=minutes - 1)
                term_df = day_df[
                    (day_df["datetime"] >= start_time)
                    & (day_df["datetime"] <= end_time)
                ]

                if len(term_df) == 0:
                    # データの終端に到達と判断
                    b_loop = False
                    continue

                # 列ごとのデータ
                open = term_df["open"].values[0]
                close = term_df["close"].values[-1]
                high = term_df["high"].max()
                low = term_df["low"].min()
                volume = term_df["volume"].sum()

                result_list.append([start_time, open, high, low, close, volume])

                # 次ループのために開始時間追加
                start_time = start_time + datetime.timedelta(minutes=minutes)

            print("end parsing day({})".format(day))

        # 結果
        result_df = pd.DataFrame(result_list)
        result_df.columns = ["datetime", "open", "high", "low", "close", "volume"]
        result_df = result_df.sort_values("datetime")

        # ファイル出力
        result_df.to_csv(os.path.join(o_dir, o_fname), index=None)

    # 注意: エクセルファイル名が西暦.xlsxになっている。& シートのデータはすでにソートしてる前提
    # 一分足専用でシート名は1minとしないといけない
    def make_all_1min_history(self, fname: str, i_dir: str, o_dir: str) -> None:
        period_name: str = "1min"

        print(i_dir, o_dir, period_name)

        # csvファイルで出力
        self._create_csv_file(os.path.join(o_dir, fname))
        # input_dirからエクセルファイルを抜き出す
        file_paths: list = self._get_data_files(input_dtr=i_dir)

        for file_path in file_paths:
            # エクセルファイルのシートからperoid_nameの名前に該当するシートを抜き出す
            wb, target_sheet_names = self._open_xls(file_path)
            # シートからデータを追加する
            self._add_csv_data(wb, target_sheet_names=target_sheet_names)

    def _get_data_files(self, input_dtr: str) -> list:
        dir_path = "input/"
        dir_list = os.listdir(dir_path)

        # データ対象のファイルのみ抽出
        # https://liquidjumper.com/programming/python/python-%E3%83%AF%E3%82%A4%E3%83%AB%E3%83%89%E3%82%AB%E3%83%BC%E3%83%89%E3%81%A7%E6%8B%A1%E5%BC%B5%E5%AD%90%E3%82%92%E9%99%90%E5%AE%9A%E3%81%97%E3%83%95%E3%82%A1%E3%82%A4%E3%83%AB%E3%82%92%E5%8F%96

        files: list = list()
        for i in range(len(dir_list)):
            if ".xlsx" == os.path.splitext(dir_list[i])[1]:
                print(os.path.join(dir_path, dir_list[i]))
                files.append(os.path.join(dir_path, dir_list[i]))
        return files

    def _create_csv_file(self, file_path: str):
        # 書き込むと一行スペースが入る問題の対処
        # https://qiita.com/ryokurta256/items/defc553f5165c88eac95
        self.csv_file = open(file_path, "w", newline="", encoding="utf-8")

        # ヘッダーを記入
        self.csv_file.write(
            ",".join(["open", "high", "low", "close", "volume", "day", "datetime"])
            + "\n"
        )

    def _add_csv_data(self, wb: openpyxl.Workbook, target_sheet_names: list):
        for sheet_name in target_sheet_names:
            print("check sheet({})".format(sheet_name))

            sheet = wb[sheet_name]

            last_row = sheet.max_row + 1
            first_row = 1
            # データ開始行がデータによってバラバラなので探す
            for i in range(first_row, last_row):
                if sheet.cell(i, 1).value == "日付":
                    first_row = i + 1
                    break

            #
            print("sheet_name({}) / max_row({})".format(sheet_name, last_row))
            # https://www.delftstack.com/ja/howto/python/python-for-loop-start-at-1/
            for i in range(first_row, last_row):
                if not sheet.cell(i, 1).value:
                    break

                # 画面に進捗状況を表示
                # https://hibiki-press.tech/python/print-progress/5416
                # しかしVSCodeのターミナルは\rの解釈がうまくいかないみたいなので、表示がバグる
                # コマンドプロンプトでの実行は表示はバグらない
                # https://ja.stackoverflow.com/questions/24834/python-3-tqdm-%E3%81%AE%E4%B8%8D%E5%85%B7%E5%90%88
                print(f"\r\033[K({i})/({last_row})", end="")

                try:
                    # 日付を取得
                    # https://www.javadrive.jp/python/date/index4.html
                    date = datetime.datetime.strptime(
                        str(sheet.cell(i, 1).value), "%Y-%m-%d %H:%M:%S"
                    )
                except Exception as ex:
                    print("{} / {}".format(ex, str(sheet.cell(i, 1).value)))
                    raise ex

                try:
                    tmp_time_date = sheet.cell(i, 2).value
                    # セルに00:00:00に記載しているデータを取得すると、正しい値が取れないバグがある(1900年のうるう年が関係しているそうだ)
                    # https://thinkami.hatenablog.com/entry/2018/12/05/220032
                    if tmp_time_date == self.excel_zero_date:
                        tmp_time_date = "00:00:00"

                    time_date = datetime.datetime.strptime(
                        str(tmp_time_date), "%H:%M:%S"
                    )
                except Exception as ex:
                    print("{} / {}".format(ex, str(sheet.cell(i, 2).value)))
                    raise ex

                # 行を追加
                row = [
                    # 始値
                    str(sheet.cell(i, 3).value),
                    # 高値
                    str(sheet.cell(i, 4).value),
                    # 安値
                    str(sheet.cell(i, 5).value),
                    # 終値
                    str(sheet.cell(i, 6).value),
                    # 出来高
                    str(sheet.cell(i, 7).value),
                    # 日にち
                    "{:04d}.{:02d}.{:02d}".format(
                        date.year,
                        date.month,
                        date.day,
                    ),
                    # 日付/時間
                    "{:04d}-{:02d}-{:02d} {:02d}:{:02d}:{:02d}".format(
                        date.year,
                        date.month,
                        date.day,
                        time_date.hour,
                        time_date.minute,
                        time_date.second,
                    ),
                ]
                self.csv_file.write(",".join(row) + "\n")

                time.sleep(0.01)

    def _open_xls(self, xls_file_path: str, peroid_name: str = "1min"):
        print(xls_file_path)
        wb = openpyxl.load_workbook(xls_file_path)

        # TODO: 該当するシート一覧を取得
        target_sheet_names: list = list()
        for sheet_name in wb.sheetnames:
            print(sheet_name)

            if sheet_name.startswith(peroid_name):
                target_sheet_names.append(sheet_name)

        print(target_sheet_names)

        return wb, target_sheet_names


def main():
    try:
        Fire(Command)
    except KeyboardInterrupt as keyex:
        # キー割込みでの例外は強制終了なのでエラーとは扱わない
        print(keyex)


if __name__ == "__main__":
    main()
